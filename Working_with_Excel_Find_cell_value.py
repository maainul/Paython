import openpyxl ,os
os.chdir('/home/mainul/Desktop/copy_files/')
print(os.getcwd())
wb=openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)
sheet=wb.get_sheet_by_name('Sheet1')
print(sheet)
for row in range(1,4):
    for column in range(1,4):
        print(sheet.cell(row=row,column=column).value)
    print('-----End of cell----')
